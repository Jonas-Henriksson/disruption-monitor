"""Load supplier data from xlsx into SQLite.

Usage:
    python -m backend.scripts.load_suppliers [--xlsx path/to/file.xlsx] [--from-s3]

The xlsx file (sometimes with .md extension) should contain columns:
    Supplier, ERP supplier desc, Supplier country, Business Area,
    Operational unit, Company country, Category L1, Category L2,
    Item description, Spend (SEK)

Run from the project root directory.
"""

from __future__ import annotations

import argparse
import os
import sqlite3
import sys
import tempfile
from collections import Counter, defaultdict
from pathlib import Path

# ── Country name normalization ────────────────────────────────
COUNTRY_ALIASES: dict[str, str] = {
    "USA": "United States",
    "U.S.A.": "United States",
    "U.S.": "United States",
    "US": "United States",
    "United States of America": "United States",
    "UK": "United Kingdom",
    "U.K.": "United Kingdom",
    "Great Britain": "United Kingdom",
    "England": "United Kingdom",
    "Brasil": "Brazil",
    "Deutschland": "Germany",
    "Bundesrepublik Deutschland": "Germany",
    "PRC": "China",
    "P.R.C.": "China",
    "Peoples Republic of China": "China",
    "People's Republic of China": "China",
    "Republic of Korea": "South Korea",
    "Korea, Republic of": "South Korea",
    "Korea": "South Korea",
    "ROC": "Taiwan",
    "Chinese Taipei": "Taiwan",
    "Czechia": "Czech Republic",
    "Czech": "Czech Republic",
    "Holland": "Netherlands",
    "The Netherlands": "Netherlands",
    "Schweiz": "Switzerland",
    "Suisse": "Switzerland",
    "Sverige": "Sweden",
    "Turkiye": "Turkey",
    "Türkiye": "Turkey",
    "Republic of Turkey": "Turkey",
    "UAE": "United Arab Emirates",
    "U.A.E.": "United Arab Emirates",
    "RSA": "South Africa",
    "Republic of South Africa": "South Africa",
    "Nippon": "Japan",
    "Bharat": "India",
    "Republic of India": "India",
    "Polska": "Poland",
    "Osterreich": "Austria",
    "Österreich": "Austria",
    "Italia": "Italy",
    "Espana": "Spain",
    "España": "Spain",
    "France (Metropolitan)": "France",
    "Russian Federation": "Russia",
    "Viet Nam": "Vietnam",
    "Viet nam": "Vietnam",
    "Srbija": "Serbia",
    "Magyarország": "Hungary",
    "Magyarorszag": "Hungary",
    "Suomi": "Finland",
    "Norge": "Norway",
    "Danmark": "Denmark",
    "Belgique": "Belgium",
    "België": "Belgium",
    "Portugal": "Portugal",
    "Hellas": "Greece",
    "Ellada": "Greece",
}

# Column name mapping: expected header -> normalized key
COLUMN_MAP = {
    "supplier": "supplier_name",
    "erp supplier desc": "erp_desc",
    "supplier country": "supplier_country",
    "business area": "business_area",
    "operational unit": "site_id",
    "company country": "company_country",
    "category l1": "category_l1",
    "category l2": "category_l2",
    "item description": "item_description",
    "spend (sek)": "spend_sek",
    "spend sek": "spend_sek",
    "spend": "spend_sek",
}


def normalize_country(name: str | None) -> str:
    """Normalize country name using alias map."""
    if not name:
        return ""
    cleaned = name.strip()
    return COUNTRY_ALIASES.get(cleaned, cleaned)


def clean_text(value) -> str:
    """Strip whitespace from text, handle None."""
    if value is None:
        return ""
    return str(value).strip()


def parse_spend(value) -> float:
    """Parse spend value to float, handling various formats."""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    # String: remove currency symbols, commas, spaces
    cleaned = str(value).replace(",", "").replace(" ", "").replace("SEK", "").replace("kr", "").strip()
    try:
        return float(cleaned)
    except (ValueError, TypeError):
        return 0.0


def download_from_s3(bucket: str, key: str) -> str:
    """Download file from S3 to temp path, return local path."""
    import boto3
    s3 = boto3.client("s3", region_name=os.environ.get("AWS_REGION", "eu-west-1"))
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    print(f"Downloading s3://{bucket}/{key} -> {tmp.name}")
    s3.download_file(bucket, key, tmp.name)
    return tmp.name


def load_markdown(path: str) -> list[dict]:
    """Parse markdown-formatted supplier data.

    Expected format per record:
        ## Supplier: {name} | Site: {site_code} {site_desc}
        - Supplier country: {country}
        - Business Area: {area}
        - Company country: {country}
        - Category: {L1} > {L2}
        - Item: {desc}
        - Spend: {amount} MSEK
    """
    import re

    records = []
    site_descriptions: dict[str, str] = {}  # site_code -> description
    current: dict | None = None

    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.rstrip("\r\n")

            # New record header
            if line.startswith("## Supplier:"):
                if current:
                    records.append(current)
                # Parse: ## Supplier: {name} | Site: {site_code} {site_desc}
                header = line[len("## Supplier:"):].strip()
                parts = header.split("|", 1)
                supplier_name = parts[0].strip()
                site_raw = parts[1].strip() if len(parts) > 1 else ""
                # Site format: "Site: 618M AEROSPACE FALCONER MU"
                site_str = site_raw[len("Site:"):].strip() if site_raw.startswith("Site:") else site_raw
                # Extract site_id (first token, e.g., "618M")
                site_parts = site_str.split(None, 1)
                site_id = site_parts[0] if site_parts else site_str
                site_desc = site_parts[1] if len(site_parts) > 1 else ""
                if site_id and site_id not in site_descriptions:
                    site_descriptions[site_id] = site_desc

                current = {
                    "supplier_name": supplier_name,
                    "site_id": site_id,
                    "supplier_country": "",
                    "business_area": "",
                    "company_country": "",
                    "category_l1": "",
                    "category_l2": "",
                    "item_description": "",
                    "spend_sek": 0.0,
                }
                continue

            if current is None:
                continue

            # Parse field lines
            if line.startswith("- Supplier country:"):
                current["supplier_country"] = line.split(":", 1)[1].strip()
            elif line.startswith("- Business Area:"):
                current["business_area"] = line.split(":", 1)[1].strip()
            elif line.startswith("- Company country:"):
                current["company_country"] = line.split(":", 1)[1].strip()
            elif line.startswith("- Category:"):
                cat = line.split(":", 1)[1].strip()
                if ">" in cat:
                    l1, l2 = cat.split(">", 1)
                    current["category_l1"] = l1.strip()
                    current["category_l2"] = l2.strip()
                else:
                    current["category_l1"] = cat
            elif line.startswith("- Item:"):
                current["item_description"] = line.split(":", 1)[1].strip()
            elif line.startswith("- Spend:"):
                spend_str = line.split(":", 1)[1].strip()
                # Parse "75.3 MSEK" -> 75_300_000 SEK
                m = re.match(r"([\d.,]+)\s*(MSEK|KSEK|SEK)?", spend_str, re.IGNORECASE)
                if m:
                    val = float(m.group(1).replace(",", ""))
                    unit = (m.group(2) or "SEK").upper()
                    if unit == "MSEK":
                        val *= 1_000_000
                    elif unit == "KSEK":
                        val *= 1_000
                    current["spend_sek"] = val

    # Don't forget the last record
    if current:
        records.append(current)

    return records, site_descriptions


def load_xlsx(path: str) -> list[dict]:
    """Parse xlsx file into list of row dicts."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("ERROR: openpyxl is required. Install with: pip install openpyxl")
        sys.exit(1)

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows()
    # Read header row
    header_row = next(rows_iter)
    headers = [clean_text(cell.value).lower() for cell in header_row]

    # Map headers to our column keys
    col_indices: dict[str, int] = {}
    for idx, header in enumerate(headers):
        for pattern, key in COLUMN_MAP.items():
            if pattern in header:
                col_indices[key] = idx
                break

    missing = {"supplier_name", "supplier_country", "site_id"} - col_indices.keys()
    if missing:
        print(f"ERROR: Missing required columns: {missing}")
        print(f"  Found headers: {headers}")
        print(f"  Mapped: {col_indices}")
        sys.exit(1)

    records = []
    for row in rows_iter:
        values = [cell.value for cell in row]
        record = {}
        for key, idx in col_indices.items():
            if idx < len(values):
                record[key] = values[idx]
            else:
                record[key] = None
        records.append(record)

    wb.close()
    return records


def detect_and_load(path: str) -> tuple[list[dict], dict[str, str]]:
    """Auto-detect file format (markdown or xlsx) and load accordingly.

    Returns (records, site_descriptions) where site_descriptions maps
    operational unit codes to their descriptions.
    """
    with open(path, "rb") as f:
        header = f.read(4)

    # xlsx files start with PK (zip signature)
    if header[:2] == b"PK":
        print("  Detected: xlsx format")
        return load_xlsx(path), {}
    else:
        print("  Detected: markdown format")
        return load_markdown(path)


def process_records(raw_records: list[dict]) -> tuple[list[tuple], dict]:
    """Process raw records into DB-ready tuples and report stats.

    Returns (db_rows, stats) where db_rows are tuples ready for executemany.
    """
    country_normalizations: Counter = Counter()
    skipped = 0
    aggregated: dict[tuple, dict] = {}

    for rec in raw_records:
        site_id = clean_text(rec.get("site_id"))
        supplier_name = clean_text(rec.get("supplier_name"))

        # Skip rows with empty required fields
        if not site_id or not supplier_name:
            skipped += 1
            continue

        # Normalize countries
        raw_country = clean_text(rec.get("supplier_country"))
        supplier_country = normalize_country(raw_country)
        if supplier_country != raw_country and raw_country:
            country_normalizations[f"{raw_country} -> {supplier_country}"] += 1

        raw_company_country = clean_text(rec.get("company_country"))
        company_country = normalize_country(raw_company_country)
        if company_country != raw_company_country and raw_company_country:
            country_normalizations[f"{raw_company_country} -> {company_country}"] += 1

        business_area = clean_text(rec.get("business_area"))
        category_l1 = clean_text(rec.get("category_l1"))
        category_l2 = clean_text(rec.get("category_l2"))
        item_description = clean_text(rec.get("item_description"))
        spend = parse_spend(rec.get("spend_sek"))

        # Composite key for aggregation
        key = (site_id, supplier_name, category_l1)
        if key in aggregated:
            aggregated[key]["spend_sek"] += spend
        else:
            aggregated[key] = {
                "site_id": site_id,
                "supplier_name": supplier_name,
                "supplier_country": supplier_country,
                "business_area": business_area,
                "company_country": company_country,
                "category_l1": category_l1,
                "category_l2": category_l2,
                "item_description": item_description,
                "spend_sek": spend,
            }

    db_rows = [
        (
            v["site_id"], v["supplier_name"], v["supplier_country"],
            v["business_area"], v["company_country"], v["category_l1"],
            v["category_l2"], v["item_description"], v["spend_sek"],
        )
        for v in aggregated.values()
    ]

    unique_suppliers = len({r[1] for r in db_rows})
    unique_sites = len({r[0] for r in db_rows})
    total_spend = sum(r[8] for r in db_rows)

    stats = {
        "total_records": len(db_rows),
        "raw_rows": len(raw_records),
        "skipped": skipped,
        "duplicates_merged": len(raw_records) - skipped - len(db_rows),
        "unique_suppliers": unique_suppliers,
        "unique_sites": unique_sites,
        "total_spend_sek": total_spend,
        "country_normalizations": dict(country_normalizations),
    }

    return db_rows, stats


def insert_into_db(db_rows: list[tuple], db_path: str | None = None,
                    site_descriptions: dict[str, str] | None = None) -> None:
    """Batch-insert rows into supplier_relationships table and site_code_map."""
    if db_path is None:
        # Use the same DB path as the app
        from backend.app.config import settings
        db_path = settings.db_path

    conn = sqlite3.connect(db_path)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")

    # Create tables if not exist
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS supplier_relationships (
            site_id         TEXT NOT NULL,
            supplier_name   TEXT NOT NULL,
            supplier_country TEXT NOT NULL,
            business_area   TEXT,
            company_country TEXT,
            category_l1     TEXT,
            category_l2     TEXT,
            item_description TEXT,
            spend_sek       REAL NOT NULL DEFAULT 0,
            PRIMARY KEY (site_id, supplier_name, category_l1)
        );
        CREATE TABLE IF NOT EXISTS site_code_map (
            site_code       TEXT PRIMARY KEY,
            site_description TEXT NOT NULL
        );
        CREATE INDEX IF NOT EXISTS idx_supplier_rel_site ON supplier_relationships(site_id);
        CREATE INDEX IF NOT EXISTS idx_supplier_rel_country ON supplier_relationships(supplier_country);
        CREATE INDEX IF NOT EXISTS idx_supplier_rel_supplier ON supplier_relationships(supplier_name);
    """)

    # Clear existing data and re-insert
    conn.execute("DELETE FROM supplier_relationships")
    conn.executemany(
        """INSERT OR REPLACE INTO supplier_relationships
           (site_id, supplier_name, supplier_country, business_area, company_country,
            category_l1, category_l2, item_description, spend_sek)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        db_rows,
    )

    # Populate site code mapping
    if site_descriptions:
        conn.execute("DELETE FROM site_code_map")
        conn.executemany(
            "INSERT OR REPLACE INTO site_code_map (site_code, site_description) VALUES (?, ?)",
            [(code, desc) for code, desc in site_descriptions.items()],
        )
        print(f"  Inserted {len(site_descriptions)} site code mappings")

    conn.commit()
    conn.close()


def main():
    parser = argparse.ArgumentParser(description="Load supplier data from xlsx into SQLite")
    parser.add_argument("--xlsx", type=str, help="Path to xlsx file")
    parser.add_argument("--from-s3", action="store_true", help="Download from S3 knowledge base")
    parser.add_argument("--db-path", type=str, default=None, help="Override DB path")
    args = parser.parse_args()

    if args.from_s3:
        bucket = "sc-monitor-frontend-317683112105"
        key = "knowledge-base/supplier_data.md"
        xlsx_path = download_from_s3(bucket, key)
    elif args.xlsx:
        xlsx_path = args.xlsx
    else:
        print("ERROR: Provide --xlsx <path> or --from-s3")
        sys.exit(1)

    print(f"Loading supplier data from: {xlsx_path}")

    raw_records, site_descriptions = detect_and_load(xlsx_path)
    print(f"  Parsed {len(raw_records)} raw rows from xlsx")

    db_rows, stats = process_records(raw_records)

    print("\n--- Processing Summary ---")
    print(f"  Raw rows:           {stats['raw_rows']}")
    print(f"  Skipped (empty):    {stats['skipped']}")
    print(f"  Duplicates merged:  {stats['duplicates_merged']}")
    print(f"  Final records:      {stats['total_records']}")
    print(f"  Unique suppliers:   {stats['unique_suppliers']}")
    print(f"  Unique sites:       {stats['unique_sites']}")
    print(f"  Total spend (SEK):  {stats['total_spend_sek']:,.0f}")

    if stats["country_normalizations"]:
        print("\n--- Country Normalizations ---")
        for mapping, count in sorted(stats["country_normalizations"].items()):
            print(f"  {mapping}: {count} rows")

    db_path = args.db_path
    insert_into_db(db_rows, db_path, site_descriptions)
    print(f"\nInserted {stats['total_records']} records into supplier_relationships table")

    # Cleanup temp file if downloaded from S3
    if args.from_s3:
        import os
        os.unlink(xlsx_path)
        print(f"Cleaned up temp file: {xlsx_path}")

    print("Done.")


if __name__ == "__main__":
    main()
